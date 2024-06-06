import os
import re
import numpy as np
import openpyxl

def analysis():
    workbook = openpyxl.load_workbook("SUSPICIOUSNESS SCORE-BASED RANKING")
    worksheet = workbook.active
    suspiciousness = {}
    finalsuspiciousness = {}
    for row in worksheet.iter_rows():
        suspiciousness[row[1].value] = row[2].value
        finalsuspiciousness[row[1].value] = row[2].value
    mrange = max(suspiciousness.values()) - min(suspiciousness.values()) + 1
    workbook = openpyxl.load_workbook("VOTE-BASED RANKING")
    worksheet = workbook.active
    votes = {}
    for row in worksheet.iter_rows():
        votes[row[0].value] = row[1].value
        finalsuspiciousness[row[0].value] = finalsuspiciousness[row[0].value] + row[1].value * mrange

    print(finalsuspiciousness)